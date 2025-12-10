# -*- coding: utf-8 -*-
import openpyxl
import sys

# UTF-8 출력 설정
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('D:/xyz/MONEY/TEST/STAFF_LIST_2025_11 (2).xlsx')
ws = wb.active

print(f"Sheet Name: {ws.title}")
print(f"Max Row: {ws.max_row}")
print(f"Max Col: {ws.max_column}")
print("\n" + "="*80)

# 처음 10행 읽기
for row_idx in range(1, min(11, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value if cell.value is not None else ''
        row_data.append(str(value))
    print(f"Row {row_idx}: {' | '.join(row_data)}")
