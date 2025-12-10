# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

# 워크북 생성
wb = openpyxl.Workbook()

# ========== 시트 1: 출퇴근 기록 (N CONG 스타일) ==========
ws1 = wb.active
ws1.title = "CHAM CONG"

# 스타일 정의
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
orange_fill = PatternFill(start_color='FFB347', end_color='FFB347', fill_type='solid')

# 제목
ws1.merge_cells('A1:AJ1')
ws1['A1'] = 'BẢNG CHẤM CÔNG - MONEY INTEGRATION TEMPLATE'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align

ws1.merge_cells('A2:AJ2')
ws1['A2'] = 'Tháng: _____ / Năm: _____'
ws1['A2'].alignment = center_align

# 헤더 행 (Row 4)
headers = ['STT', 'CODE', 'Họ Và Tên', 'Loại']
for day in range(1, 32):
    headers.append(str(day))
headers.extend(['Tổng', 'Ghi Chú'])

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    if col <= 4:
        cell.fill = yellow_fill

# 샘플 직원 데이터 (22명)
employees = [
    ('KQ-001', 'Nguyễn Văn Nam'),
    ('KQ-002', 'Nguyễn văn Hoàng'),
    ('KQ-003', 'Nguyễn văn Bình'),
    ('KQ-004', 'Nguyễn văn Ngọt'),
    ('KQ-005', 'Trương Nguyên Khôi'),
    ('KQ-006', 'Nguyễn Văn Gấp'),
    ('KQ-007', 'Phan Văn Nghĩa'),
    ('KQ-008', 'Phạm văn Xi'),
    ('KQ-009', 'Huỳnh Ngọc Sang'),
    ('KQ-010', 'Bùi văn Triệu'),
    ('KQ-011', 'Nguyễn Thị Kim Nhan'),
    ('KQ-012', 'Nguyễn Chí Trung'),
    ('KQ-013', 'Nguyễn Quốc Việt'),
    ('KQ-014', 'Nguyễn Thị Trang'),
    ('KQ-015', 'Nguyễn Ngọc Vân'),
    ('KQ-016', 'Nguyễn Phát Hải'),
    ('KQ-017', 'Huỳnh Thị Hương'),
    ('KQ-018', 'Nguyễn Văn Tâm'),
    ('KQ-019', 'Nguyễn văn Mẫn'),
    ('KQ-020', 'Trương Minh Kỳ'),
    ('KQ-021', 'Phạm Thanh Liêm'),
    ('KQ-022', 'Lê Hữu Phúc'),
]

# 근무 유형
work_types = [
    ('Giờ Chính', green_fill),      # 정상 근무
    ('Tăng Ca', orange_fill),        # 야근
    ('Ca Đêm', blue_fill),           # 야간
    ('Chủ Nhật', PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'))  # 휴일
]

row = 5
for emp_idx, (code, name) in enumerate(employees, 1):
    for type_idx, (work_type, fill) in enumerate(work_types):
        ws1.cell(row=row, column=1, value=emp_idx if type_idx == 0 else '').border = thin_border
        ws1.cell(row=row, column=2, value=code if type_idx == 0 else '').border = thin_border
        ws1.cell(row=row, column=3, value=name if type_idx == 0 else '').border = thin_border

        type_cell = ws1.cell(row=row, column=4, value=work_type)
        type_cell.border = thin_border
        type_cell.fill = fill
        type_cell.alignment = center_align

        # 날짜 셀 (빈칸 - 사용자가 입력)
        for day in range(1, 32):
            cell = ws1.cell(row=row, column=4 + day, value='')
            cell.border = thin_border
            cell.alignment = center_align

        # 합계 수식
        start_col = get_column_letter(5)
        end_col = get_column_letter(35)
        sum_cell = ws1.cell(row=row, column=36, value=f'=SUM({start_col}{row}:{end_col}{row})')
        sum_cell.border = thin_border
        sum_cell.alignment = center_align

        # 비고
        ws1.cell(row=row, column=37, value='').border = thin_border

        row += 1

# 열 너비 설정
ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 10
for col in range(5, 38):
    ws1.column_dimensions[get_column_letter(col)].width = 4

# ========== 시트 2: 간단 버전 (MONEY 직접 연동) ==========
ws2 = wb.create_sheet("MONEY_IMPORT")

# 제목
ws2.merge_cells('A1:H1')
ws2['A1'] = 'MONEY 급여계산기 연동 데이터'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

ws2['A2'] = '※ 이 시트는 MONEY 프로그램에서 직접 불러올 수 있는 형식입니다'

# 헤더
simple_headers = ['날짜', 'CODE', '이름', '정상(h)', '야근(h)', '야간(h)', '휴일(h)', '비고']
for col, header in enumerate(simple_headers, 1):
    cell = ws2.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = yellow_fill

# 샘플 데이터 (처음 2명, 며칠치)
sample_data = [
    ('2025-12-01', 'KQ-001', 'Nguyễn Văn Nam', 8, 2, 0, 0, ''),
    ('2025-12-02', 'KQ-001', 'Nguyễn Văn Nam', 8, 0, 0, 0, ''),
    ('2025-12-03', 'KQ-001', 'Nguyễn Văn Nam', 8, 1, 0, 0, ''),
    ('2025-12-01', 'KQ-002', 'Nguyễn văn Hoàng', 8, 0, 0, 0, ''),
    ('2025-12-02', 'KQ-002', 'Nguyễn văn Hoàng', 8, 3, 0, 0, ''),
    ('2025-12-07', 'KQ-001', 'Nguyễn Văn Nam', 0, 0, 0, 8, '일요일 특근'),
]

for row_idx, data in enumerate(sample_data, 5):
    for col_idx, value in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align

# 열 너비
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 20
for col in range(4, 8):
    ws2.column_dimensions[get_column_letter(col)].width = 10
ws2.column_dimensions['H'].width = 15

# ========== 시트 3: 사용 설명 ==========
ws3 = wb.create_sheet("HUONG DAN")

instructions = [
    "HƯỚNG DẪN SỬ DỤNG - 사용 설명서",
    "",
    "=== CHAM CONG 시트 (N CONG 스타일) ===",
    "1. CODE: 직원 코드 (KQ-001, KQ-002, ...)",
    "2. Họ Và Tên: 직원 이름",
    "3. Loại (유형):",
    "   - Giờ Chính: 정상 근무 시간 (8시간 기준)",
    "   - Tăng Ca: 야근 시간 (8시간 초과분)",
    "   - Ca Đêm: 야간 근무 시간",
    "   - Chủ Nhật: 일요일/휴일 근무 시간",
    "4. 1~31: 해당 날짜의 시간 입력",
    "5. Tổng: 자동 합계",
    "",
    "=== MONEY_IMPORT 시트 (간단 버전) ===",
    "1. 날짜: YYYY-MM-DD 형식 (예: 2025-12-01)",
    "2. CODE: 직원 코드",
    "3. 이름: 직원 이름",
    "4. 정상(h): 정상 근무 시간",
    "5. 야근(h): 야근 시간",
    "6. 야간(h): 야간 근무 시간",
    "7. 휴일(h): 휴일 근무 시간",
    "",
    "=== MONEY 프로그램 연동 ===",
    "1. MONEY 프로그램의 급여대장에서 '출근데이터 불러오기' 클릭",
    "2. 이 파일 선택",
    "3. 자동으로 달력에 데이터가 입력됩니다",
    "",
    "※ 주의: CODE가 일치해야 매칭됩니다!",
]

for row_idx, text in enumerate(instructions, 1):
    ws3.cell(row=row_idx, column=1, value=text)

ws3.column_dimensions['A'].width = 60

# 파일 저장
output_path = r'D:\xyz\MONEY\TEST\MONEY_TEMPLATE_2025_12.xlsx'
wb.save(output_path)
print(f'✅ 템플릿 생성 완료: {output_path}')
