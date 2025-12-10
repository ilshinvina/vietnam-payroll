# -*- coding: utf-8 -*-
import openpyxl
import sys

# UTF-8 출력 설정
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('D:/xyz/MONEY/TEST/BANG LUONG THANG 11-2025  khánh quỳnh  CU CHI ok  JS.xlsx')

print("=== 시트 목록 ===")
print(wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== 시트: {sheet_name} ===")
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    print("="*100)

    # 처음 20행 읽기, 처음 15열만
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(16, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value if cell.value is not None else ''
            # 길이 제한
            val_str = str(value)[:15]
            row_data.append(val_str)
        if any(v for v in row_data):  # 빈 행 스킵
            print(f"R{row_idx}: {' | '.join(row_data)}")
