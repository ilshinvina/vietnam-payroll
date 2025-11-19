import pandas as pd
import openpyxl
import sys

# UTF-8 출력 설정
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# 엑셀 파일 읽기
file_path = r"D:\xyz\MONEY\BANG LUONG THANG 4-20254  khánh quỳnh  CU CHI.xlsx"

# 워크북 열기
wb = openpyxl.load_workbook(file_path, data_only=True)

print("=" * 80)
print("EXCEL FILE ANALYSIS START")
print("=" * 80)

# 모든 시트 이름 출력
print(f"\nSheet List: {wb.sheetnames}\n")

# LUONG 시트 분석
if 'LUONG' in wb.sheetnames:
    sheet = wb['LUONG']

    print("=" * 80)
    print("LUONG SHEET ANALYSIS")
    print("=" * 80)

    # 헤더 행 찾기 (처음 10행 출력)
    print("\nFirst 10 Rows:\n")
    for row_idx in range(1, 11):
        row_data = []
        for col_idx in range(1, 30):  # A부터 약 30열까지
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None:
                row_data.append(f"[{col_idx}]{value}")
        if row_data:
            print(f"행 {row_idx}: {' | '.join(row_data[:15])}")  # 처음 15개만

    # 데이터 행 분석 (7행부터 ~ 20행까지)
    print("\n" + "=" * 80)
    print("Data Rows (7-15) - Column Details:")
    print("=" * 80)

    for row_idx in range(7, 16):
        print(f"\n--- 행 {row_idx} ---")
        row_has_data = False
        for col_idx in range(1, 30):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None and str(value).strip():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"  {col_letter}{row_idx} (열{col_idx}): {value}")
                row_has_data = True
        if not row_has_data:
            break

    # 헤더 확인 (4-6행)
    print("\n" + "=" * 80)
    print("Header Structure (Row 4-6):")
    print("=" * 80)
    for row_idx in range(4, 7):
        print(f"\n행 {row_idx}:")
        for col_idx in range(1, 30):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is not None and str(value).strip():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"  {col_letter}: {value}")

    # 샘플 직원 1명 완전 분석
    print("\n" + "=" * 80)
    print("First Employee Complete Analysis (Row 7):")
    print("=" * 80)
    sample_row = 7
    for col_idx in range(1, 35):
        cell = sheet.cell(row=sample_row, column=col_idx)
        value = cell.value
        col_letter = openpyxl.utils.get_column_letter(col_idx)

        # 헤더 찾기 (4-6행에서)
        header = None
        for h_row in [4, 5, 6]:
            h_cell = sheet.cell(row=h_row, column=col_idx)
            if h_cell.value:
                header = h_cell.value
                break

        if value is not None:
            print(f"  {col_letter}{sample_row} (열{col_idx:2d}) [{header}]: {value}")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
