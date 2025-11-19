import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('BANG LUONG THANG 4-20254  khánh quỳnh  CU CHI.xlsx', data_only=True)
sheet = wb['LUONG']

print('=' * 80)
print('INSURANCE RATES (Row 4-6, Column T-Z)')
print('=' * 80)

for row_idx in range(4, 7):
    print(f'\nRow {row_idx}:')
    for col_idx in range(20, 27):
        cell = sheet.cell(row=row_idx, column=col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        value = cell.value
        if value:
            print(f'  {col_letter}{row_idx}: {value}')

print('\n' + '=' * 80)
print('FIRST EMPLOYEE INSURANCE (Row 8)')
print('=' * 80)
print(f'\nBasic Salary (C8): {sheet.cell(row=8, column=3).value}')
print()

for col_idx in range(20, 28):
    cell = sheet.cell(row=8, column=col_idx)
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    value = cell.value
    print(f'  {col_letter}8: {value}')

# Calculate expected values
basic_salary = sheet.cell(row=8, column=3).value
print('\n' + '=' * 80)
print('VERIFICATION')
print('=' * 80)
print(f'Basic Salary: {basic_salary:,}')
print(f'\nEmployee Deductions:')
print(f'  W (8%): {basic_salary * 0.08:,.0f}')
print(f'  X (1.5%): {basic_salary * 0.015:,.0f}')
print(f'  Y (1%): {basic_salary * 0.01:,.0f}')
print(f'  Z (1%): {basic_salary * 0.01:,.0f}')
print(f'\nCompany Contributions:')
print(f'  T (17.5%): {basic_salary * 0.175:,.0f}')
print(f'  U (3%): {basic_salary * 0.03:,.0f}')
print(f'  V (1%): {basic_salary * 0.01:,.0f}')
